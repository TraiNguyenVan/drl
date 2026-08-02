import os
import re
import csv
import sys
import argparse
import urllib.request
import openpyxl

# All criteria in the exact order expected by the pipeline and Word template
CRITERIA_ORDER = [
    "1.1", "1.2", "1.3", "1.4", "1.5", "TC1",
    "2.1", "2.2", "2.3", "TC2",
    "3.1", "3.2", "3.3", "3.4", "3.5", "TC3",
    "4.1", "4.2", "4.3", "4.4", "4.5", "4.6", "TC4",
    "5.1", "5.2", "5.3", "TC5",
    "TOTAL"
]

CRITERIA_LEAVES = ["1.1", "1.2", "1.3", "1.4", "1.5", "2.1", "2.2", "2.3", "3.1", "3.2", "3.3", "3.4", "3.5", "4.1", "4.2", "4.3", "4.4", "4.5", "4.6", "5.1", "5.2", "5.3"]

TC_GROUPS = {
    "TC1": ["1.1", "1.2", "1.3", "1.4", "1.5"],
    "TC2": ["2.1", "2.2", "2.3"],
    "TC3": ["3.1", "3.2", "3.3", "3.4", "3.5"],
    "TC4": ["4.1", "4.2", "4.3", "4.4", "4.5", "4.6"],
    "TC5": ["5.1", "5.2", "5.3"]
}

CRITERION_NAMES = {
    "1.1": "Ý thức và thái độ trong học tập",
    "1.2": "Kết quả học tập trong kỳ học",
    "1.3": "Ý thức chấp hành tốt nội quy về các kỳ thi",
    "1.4": "Ý thức và thái độ tham gia các hoạt động ngoại khóa",
    "1.5": "Tinh thần vượt khó phấn đấu vươn lên trong học tập",
    "TC1": "Mức điểm tối đa Tiêu chí 1",
    "2.1": "Thực hiện nghiêm túc các nội quy quy chế",
    "2.2": "Thực hiện nghiêm túc các buổi họp lớp",
    "2.3": "Tham gia các buổi hội thảo việc làm",
    "TC2": "Mức điểm tối đa Tiêu chí 2",
    "3.1": "Tham gia đầy đủ các hoạt động chính trị xã hội",
    "3.2": "Tham gia công tác xã hội",
    "3.3": "Tuyên truyền tích cực hình ảnh về Trường/Khoa",
    "3.4": "Tích cực tham gia các hoạt động phòng chống tội phạm",
    "3.5": "Đưa các thông tin sai lệch thiếu tích cực về Học viện",
    "TC3": "Mức điểm tối đa Tiêu chí 3",
    "4.1": "Chấp hành nghiêm chỉnh chủ trương của Đảng",
    "4.2": "Tích cực tham gia tuyên truyền chủ trương của Đảng",
    "4.3": "Có mối quan hệ đúng mực với Thầy/ Cô",
    "4.4": "Có mối quan hệ tốt với bạn bè trong lớp",
    "4.5": "Được biểu dung khen thưởng",
    "4.6": "Vi phạm an ninh trật tự xã hội an toàn giao thông",
    "TC4": "Mức điểm tối đa Tiêu chí 4",
    "5.1": "Sinh viên làm lớp trưởng lớp phó",
    "5.2": "Thành viên tham gia các Câu lạc bộ",
    "5.3": "Sinh viên đạt thành tích đặc biệt",
    "TC5": "Mức điểm tối đa Tiêu chí 5",
    "TOTAL": "TỔNG CỘNG"
}

CRITERION_MAX = {
    "1.1": 3, "1.2": 10, "1.3": 4, "1.4": 2, "1.5": 1, "TC1": 20,
    "2.1": 15, "2.2": 5, "2.3": 5, "TC2": 25,
    "3.1": 10, "3.2": 4, "3.3": 3, "3.4": 3, "3.5": 0, "TC3": 20,
    "4.1": 8, "4.2": 5, "4.3": 5, "4.4": 5, "4.5": 2, "4.6": 0, "TC4": 25,
    "5.1": 4, "5.2": 3, "5.3": 3, "TC5": 10,
    "TOTAL": 100
}

def parse_dob(dob_str):
    if not dob_str:
        return ""
    dob_str = str(dob_str).strip()
    # Handle YYYY-MM-DD format (standard ISO format from Google Forms date field)
    match_ymd = re.search(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", dob_str)
    if match_ymd:
        year, month, day = int(match_ymd.group(1)), int(match_ymd.group(2)), int(match_ymd.group(3))
        return f"{day:02d}/{month:02d}/{year}"
    # Handle DD/MM/YYYY format
    match_dmy = re.search(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", dob_str)
    if match_dmy:
        day, month, year = int(match_dmy.group(1)), int(match_dmy.group(2)), int(match_dmy.group(3))
        return f"{day:02d}/{month:02d}/{year}"
    return dob_str

def map_headers(headers):
    # Map raw headers to student metadata and criteria codes
    mapping = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        header_lower = str(header).lower()
        if (("mã" in header_lower and "sinh viên" in header_lower) or "msv" in header_lower):
            mapping["student_id"] = idx
        elif "ngày sinh" in header_lower or "dob" in header_lower:
            mapping["dob"] = idx
        else:
            for crit in CRITERIA_LEAVES:
                # Matches patterns like "1.1" or "1.1." or "Criteria 1.1"
                if re.search(r"\b" + re.escape(crit) + r"\b", str(header)):
                    mapping[crit] = idx
                    break
    return mapping

def read_sheet_rows(sheet):
    rows = []
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
        if any(v is not None for v in row_vals):
            rows.append([str(v) if v is not None else "" for v in row_vals])
    return rows

def extract_scores_from_rows(rows):
    if not rows or len(rows) < 2:
        return {}
        
    headers = rows[0]
    mapping = map_headers(headers)
    
    if "student_id" not in mapping:
        return {}
        
    scores_by_msv = {}
    for row in rows[1:]:
        if len(row) <= max(mapping.values()):
            continue
        msv = row[mapping["student_id"]].strip().upper()
        if not msv:
            continue
            
        scores = {}
        for crit in CRITERIA_LEAVES:
            val_str = row[mapping[crit]].strip() if crit in mapping else "0"
            try:
                scores[crit] = float(val_str.replace(",", "."))
            except ValueError:
                scores[crit] = 0.0
                
        # Calculate TC totals
        for tc, children in TC_GROUPS.items():
            scores[tc] = sum(scores[child] for child in children)
        scores["TOTAL"] = sum(scores[tc] for tc in TC_GROUPS.keys())
        
        dob_val = ""
        if "dob" in mapping:
            dob_val = parse_dob(row[mapping["dob"]])
            
        scores_by_msv[msv] = {
            "dob": dob_val,
            "scores": scores
        }
    return scores_by_msv

def sync_data(xlsx_path, workspace):
    if not os.path.exists(xlsx_path):
        print(f"Error: Excel file {xlsx_path} not found.")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames
    
    sv_rows = []
    lop_rows = []
    
    for name in sheet_names:
        name_lower = name.lower()
        if "student" in name_lower or "self" in name_lower or "responses" in name_lower:
            sv_rows = read_sheet_rows(wb[name])
        elif "class" in name_lower or "lop" in name_lower or "evaluation" in name_lower:
            lop_rows = read_sheet_rows(wb[name])
            
    if not sv_rows and len(sheet_names) > 0:
        sv_rows = read_sheet_rows(wb[sheet_names[0]])
        
    sv_data = extract_scores_from_rows(sv_rows)
    lop_data = extract_scores_from_rows(lop_rows)
    
    if not sv_data:
        print("No student self-evaluation data parsed.")
        return

    # Load existing CSV database to preserve manual fields (notes, and class/advisor scores of unadjusted students)
    target_csv = os.path.join(workspace, "ai_studio_code.csv")
    existing_db = {} # msv -> crit_id -> {sv, class, advisor, note, dob}
    if os.path.exists(target_csv):
        with open(target_csv, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                msv = row["student_id"].strip().upper()
                crit_id = row["criterion_id"].strip()
                if msv not in existing_db:
                    existing_db[msv] = {}
                existing_db[msv][crit_id] = {
                    "sv": row.get("student_score", ""),
                    "lop": row.get("class_score", ""),
                    "cvht": row.get("advisor_score", ""),
                    "note": row.get("note", ""),
                    "dob": row.get("dob", "")
                }

    # Load official roster from Excel
    roster_path = os.path.join(workspace, "HV_Mau 2_Tong hop KQRL cua SV.xlsx")
    valid_students = [] # list of {tt, last_name, first_name, name, msv}
    if os.path.exists(roster_path):
        wb_roster = openpyxl.load_workbook(roster_path, data_only=True)
        sheet_roster = wb_roster.active
        for r in range(14, sheet_roster.max_row + 1):
            tt = sheet_roster.cell(r, 1).value
            last_name = sheet_roster.cell(r, 2).value
            first_name = sheet_roster.cell(r, 3).value
            msv = sheet_roster.cell(r, 4).value
            
            is_valid_tt = False
            if isinstance(tt, (int, float)):
                is_valid_tt = True
            elif isinstance(tt, str):
                try:
                    float(tt)
                    is_valid_tt = True
                except ValueError:
                    pass
                    
            if is_valid_tt and last_name and first_name and msv:
                msv_str = str(msv).strip().upper()
                if msv_str.startswith("N25"):
                    full_name = f"{str(last_name).strip()} {str(first_name).strip()}"
                    full_name = " ".join(full_name.split())
                    valid_students.append({
                        "tt": int(float(tt)),
                        "msv": msv_str,
                        "name": full_name
                    })

    # If roster could not be loaded, use keys from existing db
    if not valid_students:
        print("Warning: Roster Excel not loaded. Using existing CSV keys.")
        for msv in sorted(existing_db.keys()):
            valid_students.append({"msv": msv})

    # Rebuild database rows
    updated_rows = []
    
    for s in valid_students:
        msv = s["msv"]
        
        # Determine dob (prioritize Form responses over existing db)
        dob_val = ""
        if msv in sv_data and sv_data[msv]["dob"]:
            dob_val = sv_data[msv]["dob"]
        elif msv in existing_db:
            # Get any non-empty dob
            for c_id in CRITERIA_ORDER:
                if c_id in existing_db[msv] and existing_db[msv][c_id]["dob"]:
                    dob_val = existing_db[msv][c_id]["dob"]
                    break
        
        # Build all 28 rows in exact criteria order
        for crit_id in CRITERIA_ORDER:
            # Default to "0" (absolute 0) for students not in the Google Sheet
            sv_score = "0"
            lop_score = "0"
            cvht_score = "0"
            note = ""
            
            # Load note if student existed in database (preserving comments)
            if msv in existing_db and crit_id in existing_db[msv]:
                note = existing_db[msv][crit_id]["note"]
            
            # Overwrite with student self evaluation from Google Form
            if msv in sv_data:
                sv_score = str(sv_data[msv]["scores"].get(crit_id, 0.0)).replace(".0", "")
                
            # Overwrite with Class evaluation review sheet (also populates Advisor)
            if msv in lop_data:
                val = str(lop_data[msv]["scores"].get(crit_id, 0.0)).replace(".0", "")
                lop_score = val
                cvht_score = val
                
            updated_rows.append({
                "student_id": msv,
                "semester": "II",
                "criterion_id": crit_id,
                "criterion_name": CRITERION_NAMES[crit_id],
                "max_points": str(CRITERION_MAX[crit_id]),
                "student_score": sv_score,
                "class_score": lop_score,
                "advisor_score": cvht_score,
                "note": note,
                "dob": dob_val
            })

    # Write the rebuilt database
    fields = ["student_id", "semester", "criterion_id", "criterion_name", "max_points", "student_score", "class_score", "advisor_score", "note", "dob"]
    with open(target_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(updated_rows)
        
    print(f"Successfully synced spreadsheet scores to {target_csv}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync Google Form responses and class/advisor evaluations with database.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--sheet-id", type=str, help="Google Sheets ID of the shared evaluation spreadsheet.")
    group.add_argument("--xlsx", type=str, help="Path to local downloaded Excel spreadsheet.")
    
    args = parser.parse_args()
    workspace = os.path.dirname(os.path.abspath(__file__))
    
    if args.sheet_id:
        url = f"https://docs.google.com/spreadsheets/d/{args.sheet_id}/export?format=xlsx"
        xlsx_path = os.path.join(workspace, "temp_evaluation.xlsx")
        print(f"Downloading shared Google Sheet from: {url}")
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req) as response, open(xlsx_path, 'wb') as out_file:
                out_file.write(response.read())
            print("Download completed successfully.")
        except Exception as e:
            print(f"Error downloading Google Sheet: {e}")
            sys.exit(1)
    else:
        xlsx_path = os.path.abspath(args.xlsx)
        
    sync_data(xlsx_path, workspace)
    
    # Clean up temp file
    if args.sheet_id and os.path.exists(xlsx_path):
        os.remove(xlsx_path)
